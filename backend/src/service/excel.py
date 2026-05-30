import csv
import datetime
import io
import threading
import uuid

import openpyxl

from schema.rest import AttendeeCreate

CacheEntry = tuple[dict[str, tuple[list[str], list[list[str]]]], datetime.datetime]
_cache: dict[uuid.UUID, CacheEntry] = {}
_cache_lock = threading.Lock()
_CACHE_TTL = datetime.timedelta(minutes=30)


def _cleanup(task_id: uuid.UUID) -> None:
    with _cache_lock:
        _cache.pop(task_id, None)


def _parse_xlsx(file_bytes: bytes) -> dict[str, tuple[list[str], list[list[str]]]]:
    wb: openpyxl.Workbook | None = None
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
        result: dict[str, tuple[list[str], list[list[str]]]] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                result[sheet_name] = ([], [])
                continue
            columns = [str(c) if c is not None else "" for c in rows[0]]
            data = [[str(c) if c is not None else "" for c in row] for row in rows[1:]]
            result[sheet_name] = (columns, data)
        return result
    finally:
        if wb is not None:
            wb.close()


def _parse_csv(file_bytes: bytes) -> dict[str, tuple[list[str], list[list[str]]]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return {"Sheet1": ([], [])}
    columns = [str(c) for c in rows[0]]
    data = [[str(c) for c in row] for row in rows[1:]]
    return {"Sheet1": (columns, data)}


def parse_workbook(
    file_bytes: bytes, filename: str
) -> dict[str, tuple[list[str], list[list[str]]]]:
    name_lower = filename.lower()
    if name_lower.endswith(".csv"):
        return _parse_csv(file_bytes)
    elif name_lower.endswith(".xlsx") or name_lower.endswith(".xls"):
        return _parse_xlsx(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: {filename}")


def store_workbook(workbook: dict[str, tuple[list[str], list[list[str]]]]) -> uuid.UUID:
    task_id = uuid.uuid4()
    expiry = datetime.datetime.now() + _CACHE_TTL
    with _cache_lock:
        _cache[task_id] = (workbook, expiry)
    threading.Timer(_CACHE_TTL.total_seconds(), _cleanup, args=(task_id,)).start()
    return task_id


def get_workbook(
    task_id: uuid.UUID,
) -> dict[str, tuple[list[str], list[list[str]]]] | None:
    with _cache_lock:
        entry = _cache.get(task_id)
        if entry is None:
            return None
        workbook, expiry = entry
        if datetime.datetime.now() > expiry:
            del _cache[task_id]
            return None
        return workbook


def map_rows(
    workbook: dict[str, tuple[list[str], list[list[str]]]],
    sheet_name: str,
    title_column: str | None,
    name_column: str | None,
    raw_phone_column: str | None,
    email_column: str | None,
) -> list[AttendeeCreate]:
    if sheet_name not in workbook:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
    columns, rows = workbook[sheet_name]
    idx_by_field: dict[str, int] = {}
    if title_column is not None and title_column in columns:
        idx_by_field["title"] = columns.index(title_column)
    if name_column is not None and name_column in columns:
        idx_by_field["name"] = columns.index(name_column)
    if raw_phone_column is not None and raw_phone_column in columns:
        idx_by_field["raw_phone"] = columns.index(raw_phone_column)
    if email_column is not None and email_column in columns:
        idx_by_field["email"] = columns.index(email_column)

    def _get(row: list[str], idx: int) -> str:
        return row[idx] if idx < len(row) else ""

    result: list[AttendeeCreate] = []
    for row in rows:
        title = (
            _get(row, idx_by_field["title"]).strip() if "title" in idx_by_field else ""
        )
        name = _get(row, idx_by_field["name"]).strip() if "name" in idx_by_field else ""
        raw_phone = (
            _get(row, idx_by_field["raw_phone"]).strip()
            if "raw_phone" in idx_by_field
            else ""
        )
        email = (
            _get(row, idx_by_field["email"]).strip() if "email" in idx_by_field else ""
        )
        if not name:
            continue
        result.append(
            AttendeeCreate(title=title, name=name, email=email, raw_phone=raw_phone)
        )
    return result
